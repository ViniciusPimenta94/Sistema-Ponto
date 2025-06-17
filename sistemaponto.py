'''IMPORTANDO AS BIBLIOTECAS'''
import sys
from datetime import datetime, date
import time
import openpyxl
from PyQt5 import uic, QtWidgets
from bs4 import BeautifulSoup
import requests

class SistemaPonto:
    '''INICIANDO SISTEMA DE PONTO'''
    def __init__(self):
        # DATA E HORA ATUAL
        self.data_atual = (datetime.now()).strftime('%Y-%m-%d')
        self.hora_atual = (datetime.now()).strftime('%H:%M')

    def planilha(self):
        '''CARREGANDO PLANILHA DE PONTO'''
        user = tela.lineEdit.text()
        try:
            self.wb_ponto = openpyxl.load_workbook('D:/Executável/Ponto/Ponto Tecnologia.xlsx')
            self.ws_ponto = self.wb_ponto[user]
        except KeyError:
            print('\nUSUÁRIO NÃO IDENTIFICADO\n')
            exit()

        # QUANTIDADE DE LINHAS E COLUNAS
        self.row = self.ws_ponto.max_row
        self.column = self.ws_ponto.max_column

        # ACHANDO A LINHA DA DATA ATUAL
        i = 8
        while i <= self.row:
            dia = str(self.ws_ponto.cell(row=i, column=1).value).replace(' 00:00:00', '')
            if dia == self.data_atual:
                self.row_today = i
            i+=1

        # ENCONTRANDO AS COLUNAS
        j = 2
        while j <= self.column:
            if self.ws_ponto.cell(row=1, column=j).value == 'ENTRADA':
                self.col_entrada = j
            if self.ws_ponto.cell(row=1, column=j).value == 'SAÍDA ALMOÇO':
                self.col_saida_almoco = j
            if self.ws_ponto.cell(row=1, column=j).value == 'RETORNO ALMOÇO':
                self.col_retorno = j
            if self.ws_ponto.cell(row=1, column=j).value == 'SAÍDA':
                self.col_saida = j
            j += 1

        # MENSAGEM INICIAL
        if user == 'Amanda Mello':
            print('\nTROQUE DE MARCA DE CELULAR, IMEDIATAMENTE!!!!')
            time.sleep(1)
            print(f'\nBEM VINDA, {user.upper()}')
        else:
            print(f'\nBEM VINDO, {user.upper()}')

    def clima(self):
        '''INFORMAÇÕES SOBRE O CLIMA'''
        html = requests.get('https://www.climatempo.com.br/previsao-do-tempo/cidade/152/juizdefora-mg').content
        soup = BeautifulSoup(html, 'html.parser')

        resume = soup.find(class_='-gray -line-height-24 _center')
        temp_min = soup.find(id='min-temp-1')
        temp_max = soup.find(id='max-temp-1')

        print('\nResumo: ' + resume.text)
        print('Temperatura Mínima: ' + temp_min.string)
        print('Temperatura Máxima: ' + temp_max.string)
        time.sleep(2)

    def entrada(self):
        '''REGISTRAR ENTRADA'''
        SistemaPonto.clima(self)
        entrada = self.ws_ponto.cell(row=self.row_today, column=self.col_entrada)
        if entrada.value is None:
            entrada.value = self.hora_atual
            print(f'\nENTRADA REGISTRADA - {self.data_atual} {self.hora_atual}\n')
        SistemaPonto.save(self)

    def saida_almoco(self):
        '''REGISTRAR SAÍDA ALMOÇO'''
        saida_almoco = self.ws_ponto.cell(row=self.row_today, column=self.col_saida_almoco)
        if saida_almoco.value is None:
            saida_almoco.value = self.hora_atual
            print(f'\nSAÍDA ALMOÇO REGISTRADA - {self.data_atual} {self.hora_atual}\n')
        SistemaPonto.save(self)

    def retorno_almoco(self):
        '''REGISTRAR RETORNO ALMOÇO'''
        retorno = self.ws_ponto.cell(row=self.row_today, column=self.col_retorno)
        if retorno.value is None:
            retorno.value = self.hora_atual
            print(f'\nRETORNO ALMOÇO REGISTRADA - {self.data_atual} {self.hora_atual}\n')
        SistemaPonto.save(self)

    def saida(self):
        '''REGISTRAR SAÍDA'''
        SistemaPonto.clima(self)
        saida = self.ws_ponto.cell(row=self.row_today, column=self.col_saida)
        if saida.value is None:
            saida.value = self.hora_atual
            print(f'\nSAÍDA REGISTRADA - {self.data_atual} {self.hora_atual}\n')
        SistemaPonto.save(self)

    def hora(self):
        '''REALIZAR OPERAÇÕES COM DATETIME'''
        valor = datetime.combine(date.today(), self)
        return valor

    def save(self):
        '''SALVAR A PLANILHA'''
        i = 9
        while i <= self.row:
            horas = self.ws_ponto.cell(row=i, column=6)
            lista = ['SABADO', 'DOMINGO', 'FOLGA', 'FERIADO', 'FERIAS']
            if not horas.value in lista:
                entrada = self.ws_ponto.cell(row=i, column=self.col_entrada).value
                if isinstance(entrada, str):
                    entrada = datetime.strptime(self.ws_ponto.cell(row=i, column=self.col_entrada).value, '%H:%M').time()

                saida_almoco = self.ws_ponto.cell(row=i, column=self.col_saida_almoco).value
                if isinstance(saida_almoco, str):
                    saida_almoco = datetime.strptime(self.ws_ponto.cell(row=i, column=self.col_saida_almoco).value, '%H:%M').time()

                retorno = self.ws_ponto.cell(row=i, column=self.col_retorno).value
                if isinstance(retorno, str):
                    retorno = datetime.strptime(self.ws_ponto.cell(row=i, column=self.col_retorno).value, '%H:%M').time()

                saida = self.ws_ponto.cell(row=i, column=self.col_saida).value
                if isinstance(saida, str):
                    saida = datetime.strptime(self.ws_ponto.cell(row=i, column=self.col_saida).value, '%H:%M').time()

                if not entrada is None and not saida_almoco is None and not retorno is None and not saida is None:
                    horas.value = (SistemaPonto.hora(saida_almoco)-SistemaPonto.hora(entrada)) + (SistemaPonto.hora(saida)-SistemaPonto.hora(retorno))
            i+=1
        print('\nCONCLUÍDO')
        self.wb_ponto.save('D:/Executável/Ponto/Ponto Tecnologia.xlsx')

        if not self.ws_ponto.cell(row=self.row_today, column=self.col_saida).value is None:
            self.wb_ponto = openpyxl.load_workbook('D:/Executável/Ponto/Ponto Tecnologia.xlsx')
            self.wb_ponto.save('D:/Executável/Ponto/backup/Ponto Tecnologia.xlsx')
        exit()

app = QtWidgets.QApplication(sys.argv)
tela = uic.loadUi('D:/Executável/Ponto/assets/Interface_PONTO.ui')

tela.show()

users = ['Amanda Mello', 'Guilherme Corradi', 'Paulo Pimenta']
completer = QtWidgets.QCompleter(users)
tela.lineEdit.setCompleter(completer)

load_active = SistemaPonto()
tela.pushButton_5.clicked.connect(load_active.planilha)
tela.pushButton.clicked.connect(load_active.entrada)
tela.pushButton_2.clicked.connect(load_active.saida_almoco)
tela.pushButton_3.clicked.connect(load_active.retorno_almoco)
tela.pushButton_4.clicked.connect(load_active.saida)

sys.exit(app.exec_())
